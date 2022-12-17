from openpyxl import Workbook


wb = Workbook()
ws = wb.active
ws['A4'].value

import os

counter = 1


def write_to_xlsx(title, kitchenType, cookingTime, PortionCounter, CaloricContent,
                  Squirrels, Fats, Carbohydrates, Water, ingredients, steps,
                  category, photo_link, quote, origin):
    global counter, ws
    data = "data.xlsx"
    ws[f"A{counter}"].value = counter
    ws[f"B{counter}"].value = title
    ws[f"C{counter}"].value = kitchenType
    ws[f"D{counter}"].value = cookingTime
    ws[f"E{counter}"].value = PortionCounter
    ws[f"F{counter}"].value = CaloricContent
    ws[f"G{counter}"].value = Squirrels
    ws[f"H{counter}"].value = Fats
    ws[f"I{counter}"].value = Carbohydrates
    ws[f"J{counter}"].value = Water
    ws[f"K{counter}"].value = ingredients
    ws[f"L{counter}"].value = steps
    ws[f"M{counter}"].value = category
    ws[f"N{counter}"].value = photo_link
    ws[f"O{counter}"].value = quote
    ws[f"P{counter}"].value = origin
    wb.save(data)
    counter+=1


for address, dirs, files in os.walk(r'C:\Users\allelleo\Desktop\eat\eatsite\Test\LinkParser\kulinarenok\recipes'):
    for name in files:
        path = os.path.join(address, name)
        file = open(path, 'r')
        data = file.read()
        title = data.split("\n")[0]
        kitchenType = data.split("\n")[1]
        cookingTime = data.split("\n")[2]
        PortionCounter = data.split("\n")[3]
        CaloricContent = data.split("\n")[4]
        Squirrels = data.split("\n")[5]
        Fats = data.split("\n")[6]
        Carbohydrates = data.split("\n")[7]
        Water = data.split("\n")[8]
        ingredients = data.split("ING--#@")[1]
        steps = data.split("STEPS--&%")[1]
        category = data.split("CAT--&%")[1]
        photo_link = data.split("PH--&%")[1]
        quote = data.split("QUOE--&%")[1]
        origin = data.split("\n")[-1]
        write_to_xlsx(title, kitchenType, cookingTime, PortionCounter, CaloricContent,
                      Squirrels, Fats, Carbohydrates, Water, ingredients, steps,
                      category, photo_link, quote, origin)
        #break
    #break
